# thu vien gui
import PySimpleGUI as sg
# thu vien lay du lieu serial
import time
import serial
import serial.tools.list_ports
# thu vien chinh sua file excel
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import Workbook
from openpyxl.chart import (
    legend,
    trendline,
    marker,
    ScatterChart,
    Reference,
    Series,
)
# thu vien hinh anh
from PIL import Image, ImageTk
# thu vien he thong 
import os
import datetime
import base64
# thu vien ve do thi
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
# This is to include a matplotlib figure in a Tkinter canvas
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from sklearn.metrics import r2_score

# ham chinh sua du lieu tren bang
def edit_cell(window, key, row, col, tdata, justify='left'):

    global textvariable, edit

    def return_callback(event):
        global edit
        # event.widget gives you the same entry widget we created earlier
        widget = event.widget
        # Destroy the entry widget
        widget.destroy()
        # Destroy all widgets
        widget.master.destroy()
        edit = False
        return

    def callback(event, row, col, text, key, tdata):
        global edit, data
        # event.widget gives you the same entry widget we created earlier
        widget = event.widget
        if key == 'return':
            # Get new text that has been typed into widget
            text = widget.get()
            # Print to terminal
            print(f"col {col}, row {row-1} edited: {text}")
            # editing the data itself
            text = float(text)
            tdata[row-1][col] = text
            # re-evaluate data after changes
            tdata = evaluate_data(tdata)
            # print out edited data
            print(tdata[row-1])
            # update table
            window['-t-'].update(values=data)   

        # Destroy the entry widget
        widget.destroy()
        # Destroy all widgets
        widget.master.destroy()
        # Get the row from the table that was edited
        # table variable exists here because it was called before the callback
        values = list(table.item(row, 'values'))
        # Store new value in the appropriate row and column
        values[col] = text
        table.item(row, values=values)
        edit = False

    if edit or row <= 0:
        return

    edit = True
    # Get the Tkinter functionality for our window
    root = window.TKroot
    # Gets the Widget object from the PySimpleGUI table - a PySimpleGUI table is really
    # what's called a TreeView widget in TKinter
    table = window[key].Widget
    # Get the row as a dict using .item function and get individual value using [col]
    # Get currently selected value
    text = table.item(row, "values")[col]
    # Return x and y position of cell as well as width and height (in TreeView widget)
    x, y, width, height = table.bbox(row, col)
    wheight = int(window.size[1] / 2) + 20
    yy = wheight + y

    # Create a new container that acts as container for the editable text input widget
    frame = sg.tk.Frame(root)
    # put frame in same location as selected cell
    frame.place(x=x, y=yy, anchor="nw", width=width, height=height)

    # textvariable represents a text value
    textvariable = sg.tk.StringVar()
    textvariable.set(text)
    # Used to acceot single line text input from user - editable text input
    # frame is the parent window, textvariable is the initial value, justify is the position
    entry = sg.tk.Entry(frame, textvariable=textvariable, justify=justify)
    # Organizes widgets into blocks before putting them into the parent
    entry.pack()
    # selects all text in the entry input widget
    entry.select_range(0, sg.tk.END)
    # Puts cursor at end of input text
    entry.icursor(sg.tk.END)
    # Forces focus on the entry widget (actually when the user clicks because this initiates all this Tkinter stuff, e
    # ending with a focus on what has been created)
    entry.focus_force()
    # lambda e generates an empty function, which is turned into an event function 
    # which corresponds to the "return" (the event created when the user press enter)
    entry.bind("<Return>", lambda e, r=row, c=col, t=text, k='return', td=tdata:callback(e, r, c, t, k, td))
    # We want the user to be able to not make any accidental changes to the table data
    # so it is expected for the user to hit esc or click outside the editing area to 
    # cancel.
    entry.bind("<FocusOut>", lambda e: return_callback(e))
    entry.bind("<Escape>", lambda e: return_callback(e))

# Embedding the Matplotlib toolbar into your application
def draw_figure_w_toolbar(canvas, fig, canvas_toolbar):
    if canvas.children:
        for child in canvas.winfo_children():
            child.destroy()
    if canvas_toolbar.children:
        for child in canvas_toolbar.winfo_children():
            child.destroy()
    figure_canvas_agg = FigureCanvasTkAgg(fig, master=canvas)
    figure_canvas_agg.draw()
    toolbar = Toolbar(figure_canvas_agg, canvas_toolbar)
    toolbar.update()
    figure_canvas_agg.get_tk_widget().pack(side='right', fill='both', expand=1)


class Toolbar(NavigationToolbar2Tk):
    def __init__(self, *args, **kwargs):
        super(Toolbar, self).__init__(*args, **kwargs)

# add point to existing scatter chart
def addPoint(scat, new_point, c='k'):
    old_off = scat.get_offsets()
    new_off = np.concatenate([old_off,np.array(new_point, ndmin=2)])
    old_c = scat.get_facecolors()
    new_c = np.concatenate([old_c, np.array(mcolors.to_rgba(c), ndmin=2)])

    scat.set_offsets(new_off)
    scat.set_facecolors(new_c)

    scat.axes.figure.canvas.draw_idle()

def update_trendline(plt, plotname, x1, y1):
    z = np.polyfit(x1, y1, 1)
    p = np.poly1d(z)

    # print(f"plot {plotname}: \n", x1, "\n", x2, "\n", z)
    
    plt.figure(plotname)

    plt.plot(x1, p(x1), "r--")
    text = f"$y={z[0]:0.3f}\;x{z[1]:+0.3f}$"
    plt.gca().text(0.05, 0.95, text,transform=plt.gca().transAxes,
        fontsize=14, verticalalignment='top')

# xuat do thi excel
def export_to_excel(path, headers, data, fig1_title, fig1_headers, fig1_data, fig2_title, fig2_headers, fig2_data):

    # create a workbook
    wb = Workbook()
    ws = wb.active

    # main data
    rows = []
    rows.append(headers)
    for d in data: rows.append(d)

    # data for creating fig1
    r1 = []
    r1.append(fig1_headers)
    for d in fig1_data: r1.append(d)

    # data for creating fig2
    r2 = []
    r2.append(fig2_headers)
    for d in fig2_data: r2.append(d)

    
    for r in rows: ws.append(r)
    ws.append([])
    for r in r1: ws.append(r)
    ws.append([])
    for r in r2: ws.append(r)
    
    # create figure 1
    chart1 = ScatterChart()
    chart1.title = fig1_title
    chart1.style = 13
    chart1.x_axis.title = fig1_headers[0]
    chart1.y_axis.title = fig1_headers[1]
    chart1.legend = None

    x1value = Reference(ws, min_col=2, min_row=(len(rows) + 3), max_row=(len(rows) + 1)+len(r1))
    y1value = Reference(ws, min_col=1, min_row=(len(rows) + 3), max_row=(len(rows) + 1)+len(r1))
    series1 = Series(x1value, y1value)
    series1.marker= marker.Marker('circle')
    series1.graphicalProperties.line.noFill=True
    series1.trendline = trendline.Trendline(trendlineType='linear', dispEq=True)
    chart1.series.append(series1)

    # create figure 2
    chart2 = ScatterChart()
    chart2.title = fig2_title
    chart2.style = 13
    chart2.x_axis.title = fig2_headers[0]
    chart2.y_axis.title = fig2_headers[1]
    chart2.legend = None

    x2value = Reference(ws, min_col=2, min_row=(len(rows) + 2)+len(r1)+2, max_row=(len(rows) + 2)+len(r1)+len(r2))
    y2value = Reference(ws, min_col=1, min_row=(len(rows) + 2)+len(r1)+2, max_row=(len(rows) + 2)+len(r1)+len(r2))
    series2 = Series(x2value, y2value)
    series2.marker= marker.Marker('circle')
    series2.graphicalProperties.line.noFill=True
    series2.trendline = trendline.Trendline(trendlineType='linear', dispEq=True)
    chart2.series.append(series2)

    ws.add_chart(chart1, "I1")
    ws.add_chart(chart2, "I18")

    # tô màu
    for cell in ws["1:1"]:
        cell.fill = PatternFill(start_color='35b1de', end_color='35b1de', fill_type='solid')
        cell.font = Font(bold=True)
    
    for i in range(1, len(rows)):
        for cell in ws[f"{i+1}:{i+1}"]:
            cell.fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
            cell.font = Font(bold=False)

    wb.save(path)

# trinh chon cong com
def portselector():
    # hien thi tat ca cong serial dang mo tren may tinh
    ports = serial.tools.list_ports.comports()
    ports = sorted(ports)

    portlist = []
    for i in range(len(ports)):
        port = ports[i].name
        desc = ports[i].description
        hwid = ports[i].hwid
        portlist.append("{}. {}: {} [{}] \n".format(i+1, port, desc, hwid))
        # print("{}: {} [{}]".format(port, desc, hwid))

    # print(portlist)

    layout = [
        [sg.Text("Chọn cổng COM đến ESP32: ", background_color='#eeeeee', text_color='#000')],
        [sg.Combo(values=portlist, expand_x=True, background_color='#eeeeee', text_color='#000', button_background_color='#eeeeee', button_arrow_color="#000")],
        [sg.Submit(button_text="Kết nối", button_color=('#fff', '#000'))]
    ]
    win = sg.Window("Chọn cổng COM", layout, finalize=True, background_color='#eeeeee', font=("Arial", 10))
    e, v = win.read()
    win.close()

    # chon cong com den esp32
    # i = int(input("Chọn cổng COM để kết nối đến ESP32: "))
    port = ports[portlist.index(v[0])]
    ser = serial.Serial(str(port.name), 9600, timeout=0.050)
    return ser, str(port.description)

# ham su li du lieu vao tu thiet bi do
def datain(ser, testcount, data):
    global hl, hl2
    sout = ser.readline()
    sout_decoded = str(sout).split(";")
    print(sout_decoded)
    try:
        print(testcount, ". ", sout_decoded[2], "(ms)")
        while True:
            # tạo input box trống -> đặt key -> update nội dung theo key
            # dùng hàm ngoài để parse công thức
            layout = [
                [sg.Text(f"Nhập dữ liệu còn thiếu của lần đo thứ {testcount+1}:",  background_color='#eeeeee', text_color='#000')],
                [sg.Text(f"Dữ liệu thời gian từ bộ đo: {sout_decoded[2]}(ms)",  background_color='#eeeeee', text_color='#000')],
                [sg.Text("Lực kéo (Lý thuyết): ",  background_color='#eeeeee', text_color='#000'), sg.InputText( background_color='#fff', text_color='#000', border_width=0)],
                [sg.Text("Khối lượng:             ",  background_color='#eeeeee', text_color='#000'), sg.InputText( background_color='#fff', text_color='#000', border_width=0)],
                [sg.Text("Quãng đường:         ",  background_color='#eeeeee', text_color='#000'), sg.InputText( background_color='#fff', text_color='#000', border_width=0)],
                [sg.Submit(button_text="Hoàn tất",  button_color=('#fff', '#000'), bind_return_key=True)]
            ]
            win = sg.Window("Nhập dữ liệu đo", layout, finalize=True, background_color='#eeeeee', font=('Arial', 14), keep_on_top=True)
            e, v = win.read()
            win.close()
            if (v[0] != "" and v[1] != "" and v[2] != ""): 
                testcount += 1
                break
            else: sg.Popup("Các ô dữ liệu không được để trống!", title="Chú ý", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))
        time = float(sout_decoded[2]) / 1000 #ms to s
        acceleration = round(((float(v[2])*2)/(time*time)), 2)
        data.append([testcount, float(v[0]), float(v[1]), time, float(v[2]), acceleration])
        print(data)

        return data, testcount
    except Exception as e:
        print("oops ", e)
        sg.Popup(e, title="Lỗi", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))
        testcount =  (testcount - 1) if testcount >= 2 else testcount
        return data, testcount

# ham tinh lai ket qua 
def evaluate_data(data):
    for d in data:
        acceleration = round(((float(d[4])*2)/(d[3]*d[3])), 2)
        d[5] = acceleration
    return d

if __name__ == "__main__":
    exp_temp = {
        "name": "Định luật 2 newton",
        "img": "assets//img1.png",
        "table_data": {
            "headers": ["Lần thử", "Lực kéo (lt)", "Khối lượng", "Thời gian", "Quãng đường", "Gia tốc"],
        },
        "plot": {
            "name": "Đồ thị m+M (const)",
            "x_name": "F (N)",
            "y_name": "a (m/s)",
            "headers": ["Lực kéo (F)", "Gia tốc (m/s)"]
        },
        "plot2": {
            "name": "Đồ thị F (const)",
            "x_name": "1/(m+M) (kg)",
            "y_name": "a (m/s)",
            "headers": ["1/m+M (kg)", "Gia tốc (m/s)"],
        },
    }
    # init serial port
    ser, ser_desc = portselector()
    print(ser.name, ser_desc)

    # trang thai bang (cho sua hay khong)
    editState = False

    # bien dem so lan thu
    testcount = 0

    # trang thai bang
    edit = False

    # du lieu cua bang
    # for testing purposes - removed in release
    data = [[1, 1.0, 0.3, 0.55, 0.5, 3.31], [2, 1.0, 0.4, 0.64, 0.5, 2.44], [3, 1.0, 0.5, 0.71, 0.5, 1.98], [4, 2.0, 0.5, 0.5, 0.5, 4.0], [5, 3.0, 0.5, 0.42, 0.5, 5.67]]
    # data = []
    headings = exp_temp["table_data"]["headers"]

    # du lieu de ve do thi 1 (x: luc keo - y: gia toc)
    x = []
    y = []

    # du lieu de ve do thi 2 (x: 1/m+M - y: gia toc)
    x2 = []
    y2 = []

    # offset + sampling data
    # sample1: lọc giá trị khối lượng (m+M)
    # sample2: lọc giá trị lực (F)
    sample1 = 0.5
    sample2 = 1
    offset = 0.05

    # tao cua so chuong trinh
    sg.set_options(dpi_awareness=True)
    menu = [
        ['&Tệp', ['&Xuất đồ thị...', '&Thoát']],
        ['&Số liệu', ['&Bảng số liệu', ['Xóa bảng', 'Xóa dòng', 'Bật chỉnh sửa'], '&Đồ thị', ['Thiết lập...']]],
        ['&Trợ giúp', ['&Thông tin']]
    ]
    tab_table = [
        # bang so lieu
        # thêm khung/màu so le
        [sg.Table(
            values=data, 
            headings=headings, 
            key="-t-", 
            # auto_size_Frames=False,  
            def_col_width=10, 
            num_rows=10, 
            justification="center", 
            expand_x=True, 
            expand_y=True, 
            font=("Arial", 14, "bold"), 
            # header_background_color=(), header_text_color=(),
            alternating_row_color = "#add8e6",
            selected_row_colors = ("#000", "#86a8b3"),
            header_relief=sg.RELIEF_SOLID,
            background_color='#fff', 
            text_color='#000',
            sbar_trough_color='#fff', 
            sbar_background_color='#eeeeee', 
            sbar_arrow_color='#fff', 
            sbar_frame_color='#eeeeee', 
            sbar_relief=sg.RELIEF_FLAT,
            enable_click_events=True
        )],
    ]
    tab_plot1 = [
         # do thi bieu dien ket qua thi nghiem
        [sg.Canvas(key='controls_cv', background_color='#eeeeee')],
        [sg.Column(
            layout=[
                [sg.Canvas(key='fig_cv', expand_x=True, expand_y=True, size=(400 * 2, 400))]
                ],
            background_color='#eeeeee', pad=(0, 0), expand_x=True, expand_y=True),
        ],
    ]
    tab_plot2 = [
         # do thi bieu dien ket qua thi nghiem
        [sg.Canvas(key='controls_cv2', background_color='#eeeeee')],
        [sg.Column(
            layout=[
                [sg.Canvas(key='fig_cv2', expand_x=True, expand_y=True, size=(400 * 2, 400))]
                ],
            background_color='#eeeeee', pad=(0, 0), expand_x=True, expand_y=True),
        ],
    ]
    layout = [
        [sg.Menu(menu, tearoff=False, key='-menu-')],
        [
            sg.Column([
                [
                    # hinh anh thi nghiem
                    sg.Frame("Thiết lập thí nghiệm", [
                        [sg.Image(expand_x=True, expand_y=True, key="-img-", background_color='#eeeeee')]
                    ], background_color='#eeeeee', title_color="#000", key="-2-"),
                ],
                [
                    # bảng số liệu
                    sg.Frame("Bảng số liệu", tab_table, background_color='#eeeeee', title_color="#000", key="-1-"),
                ],
            ], background_color='#eeeeee'),
            sg.Frame("Kết quả", [
                [sg.TabGroup([
                    [
                        # sg.Tab('Bảng số liệu', tab_table, background_color='#eeeeee'),
                        sg.Tab('Đồ thị 1', tab_plot1, background_color='#eeeeee', key='tab1'),
                        sg.Tab('Đồ thị 2', tab_plot2, background_color='#eeeeee', key='tab2'),
                    ]
                ], background_color="#eeeeee", selected_title_color="#eeeee1", key="-4-", enable_events=True)],
            ], background_color='#eeeeee', title_color="#000", key="-3-")
        ],
    ]
    # tao cua so chuong trinh chinh
    win = sg.Window(f"Kết quả đo - {ser.name}: {ser_desc}", layout, resizable=True, finalize=True, background_color='#eeeeee', size=(1400, 600))
    win.bind('<Configure>', "Configure")

    # load anh lan dau
    path = exp_temp["img"]
    im = Image.open(path)
    im = im.resize((int(win.size[0] / 2), int(win.size[1] / 2)), resample=Image.BICUBIC)
    image = ImageTk.PhotoImage(image=im)
    win["-img-"].update(data=image)

    prev_tab = None

    # its apperance here is because there are a lot of varibales first declared in this condition 
    # that is in use in this function, so that's why we "can't" move it to the top port of the code
    # p/s: we could but it super annoying
    def createFig(plotname, plotnum, x, y, figure_cv, ctrl_cv):
        # ve lai do thi 1
        plt.figure(plotnum)

        plt.clf()
        fig = plt.gcf()
        DPI = fig.get_dpi()

        ax = plt.gca()
        ax.set_xlim(xmin = 0)
        ax.set_ylim(ymin = 0)

        hl = plt.scatter(x, y, cmap=matplotlib.cm.spring)
        plt.title(exp_temp[plotname]["name"])
        plt.xlabel(exp_temp[plotname]["x_name"])
        plt.ylabel(exp_temp[plotname]["y_name"])
        plt.grid(alpha=0.5)
        # plt.xticks(np.arange(1, 7, 1.0))
        # plt.yticks(np.arange(1, 7, 1.0))
        draw_figure_w_toolbar(win[figure_cv].TKCanvas, fig, win[ctrl_cv].TKCanvas)
        plt.autoscale(True)
        plt.draw()

        # ve trendline do thi 1
        if (len(x) != 0 and len(y) != 0): 
            update_trendline(plt, plotnum, x, y)
            plt.figure(plotnum)
            plt.draw()

    createFig("plot", 1, x, y, 'fig_cv', 'controls_cv')
    createFig("plot2", 2, x2, y2, 'fig_cv2', 'controls_cv2')

    # event loop
    while True:
        try:
            if (ser.in_waiting != 0):
                data, testcount = datain(ser, testcount, data)
                print(data)
                # cap nhat bang
                win["-t-"].update(values=data)
        except serial.serialutil.SerialException:
            sg.Popup("Thiết bị đo đã ngắt kết nối!", title="Thông báo", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))
            break
        e, v = win.read(timeout=500)

        # Checks if the event object is of tuple data type, indicating a click on a cell'
        if isinstance(e, tuple):
            if isinstance(e[2][0], int) and e[2][0] > -1 and editState:
                cell = row, col = e[2]
                edit_cell(win, '-t-', row+1, col, data, justify='right')

        # cap nhat do thi
        if prev_tab != v['-4-']:
            prev_tab = v['-4-']
            if v['-4-'] == 'tab1':
                print("recalulating fig 1")
                # du lieu de ve do thi 1 (x: luc keo - y: gia toc)
                x = []
                y = []
                # do thi 1 (loc du lieu co m+M = sample1 va co sai so la offset) - default la 0.5 - offset 0.05 
                for d in data:
                    if (d[2] <= (sample1+offset)) and (d[2] >= (sample1-offset)): 
                        x.append(d[1])
                        y.append(d[5])
                createFig("plot", 1, x, y, 'fig_cv', 'controls_cv')
                print(x)
                print(y)

            if v['-4-'] == 'tab2': 
                print("recalulating fig 2")
                # du lieu de ve do thi 2 (x: 1/m+M - y: gia toc)
                x2 = []
                y2 = []
                # do thi 2 (loc du lieu co F = sample2 va co sai so la offset) - default la 1 - offset 0.05
                for d in data:
                    if (d[1] <= (sample2+offset)) and (d[1] >= (sample2-offset)): 
                        x2.append(1/float(d[2]))
                        y2.append(d[5])
                createFig("plot2", 2, x2, y2, 'fig_cv2', 'controls_cv2')
                print(x2)
                print(y2)

        if e == 'Configure':
            # lay chieu dai chieu rong
            wlength = int(win.size[0] / 2) - 20
            wheight = int(win.size[1] / 2) - 20
            # cap nhat kich thuoc cua so
            win["-1-"].set_size((wlength, wheight))
            win["-2-"].set_size((wlength, wheight))
            win["-3-"].set_size((min(win.size), min(win.size)))
            win["-4-"].set_size((min(win.size), min(win.size)))
            # cap nhat kich thuoc hinh anh
            img_length = int(wlength) - int(int(wlength) * 5 / 100)
            img_height = int(wheight) - int(int(wheight) * 10 / 100)
            im = Image.open(path)
            im = im.resize((img_length, img_height), resample=Image.BICUBIC)
            image = ImageTk.PhotoImage(image=im)
            win["-img-"].update(data=image)
        
        if e == "Xóa bảng":
            # reinit the data array
            data = []
            # reninit counter
            testcount = 0
            # updating the table w/ new data
            win['-t-'].update(values=data)
            # these's no need to update the figure as it's already been updated 
            # everytime we use it

        if e == "Xóa dòng":
            cfg_layout = [
                [
                    sg.Text("Dòng: ", background_color='#eeeeee', text_color='#000'), 
                    sg.In(background_color='#fff', text_color='#000', border_width=0)
                ],
                [sg.Button("Hoàn tất", key="-cfg_done-", button_color=('#fff', '#000'), bind_return_key=True)]
            ]
            cfg = sg.Window("Xóa", cfg_layout, element_justification='left', background_color='#eeeeee', font=("Arial", 10), finalize=True)
            cfge, cfgv = cfg.read()
            if cfge == "-cfg_done-":
                data.pop(int(cfgv[0])-1)
                win['-t-'].update(values=data)
                cfg.close()
                if (v['-4-'] == 'tab1'):
                    createFig("plot", 1, x, y, 'fig_cv', 'controls_cv')
                elif (v['-4-'] == 'tab2'):
                    createFig("plot2", 2, x2, y2, 'fig_cv2', 'controls_cv2')
            else:
                pass

        if e == "Thiết lập...":
            cfg_layout = [
                [
                    sg.Text("Lọc đồ thị 1", background_color='#eeeeee', text_color='#000'), 
                    sg.In(default_text = str(sample1), background_color='#fff', text_color='#000', border_width=0)
                ],
                [
                    sg.Text("Lọc đồ thị 2", background_color='#eeeeee', text_color='#000'), 
                    sg.In(default_text = str(sample2), background_color='#fff', text_color='#000', border_width=0)
                ],
                [
                    sg.Push(background_color='#eeeeee'), sg.Text("Sai số", background_color='#eeeeee', text_color='#000'), 
                    sg.In(default_text = str(offset), background_color='#fff', text_color='#000', border_width=0)
                ],
                [sg.Button("Hoàn tất", key="-cfg_done-", button_color=('#fff', '#000'), bind_return_key=True)]
            ]
            cfg = sg.Window("Thiết lập đồ thị", cfg_layout, element_justification='left', background_color='#eeeeee', font=("Arial", 10), finalize=True)
            cfge, cfgv = cfg.read()
            if cfge == "-cfg_done-":
                print(cfgv)
                sample1 = float(cfgv[0])
                sample2 = float(cfgv[1])
                offset = float(cfgv[2])
                cfg.close()
                if (v['-4-'] == 'tab1'):
                    createFig("plot", 1, x, y, 'fig_cv', 'controls_cv')
                elif (v['-4-'] == 'tab2'):
                    createFig("plot2", 2, x2, y2, 'fig_cv2', 'controls_cv2')
            else:
                pass
        
        if e == "Bật chỉnh sửa":
            editState = True
            menu[1][1][1][2] = "Tắt chỉnh sửa"
            win['-menu-'].update(menu_definition=menu)
        if e == "Tắt chỉnh sửa":
            editState = False
            menu[1][1][1][2] = "Bật chỉnh sửa"
            win['-menu-'].update(menu_definition=menu)

        if e == sg.WIN_CLOSED or e == "Thoát":
            break
        if e == "Xuất đồ thị...":
            dir = str(os.getcwd())
            name = datetime.datetime.now()
            name = name.strftime("%d%m%y_%H%M%S") + ".xlsx"
            layout = [
                [
                    sg.Text("Chọn thư mục: ", background_color='#eeeeee', text_color='#000'), 
                    sg.Input(key="-IN2-" ,change_submits=True, default_text=dir, background_color='#fff', text_color='#000', border_width=0), 
                    sg.FolderBrowse(key="-IN-", button_color=('#fff', '#000'))
                ],
                [
                    sg.Button("Chọn", key="Submit", button_color=('#fff', '#000'))
                ]
            ]
            exp_win = sg.Window("Xuất đồ thị", layout, finalize=True, background_color='#eeeeee')
            while True:
                event, values = exp_win.read()
                if event == sg.WIN_CLOSED or event=="Exit":
                    break
                elif event == "Submit":
                    dir = values["-IN2-"]
                    dir = dir + f"/{name}"
                    fig1_data = []
                    for i in range(len(x)):
                        fig1_data.append([x[i], y[i]])
                    fig2_data = []
                    for i in range(len(x2)):
                        fig2_data.append([x2[i], y2[i]])
                    export_to_excel(dir, exp_temp["table_data"]["headers"], data, exp_temp["plot"]["name"], exp_temp["plot"]["headers"], fig1_data, exp_temp["plot2"]["name"], exp_temp["plot2"]["headers"], fig2_data)
                    sg.Popup(f"Đã xuất {name} tại đường dẫn {dir}.", title="Hoàn tất", background_color='#eeeeee', text_color='#000', button_color=('#fff', '#000'))
                    break
            exp_win.close()

    win.close()