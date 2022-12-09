from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

wb = Workbook()
ws = wb.active

rows = [
    ["Lần thử", "Lực kéo (lt)", "Khối lượng", "Thời gian", "Quãng đường", "Gia tốc", "Lực kéo (tt)"],
]

for d in data: 
    rows.append(d)

for row in rows:
    ws.append(row)

for cell in ws["1:1"]:
    cell.fill = PatternFill(start_color='35b1de', end_color='35b1de', fill_type='solid')
    cell.font = Font(bold=True)

for rows in ws.iter_rows(min_row=2, min_col=1, max_col=7, max_row=len(rows)):
    for cell in rows:
        cell.fill = PatternFill(start_color='35de8c', end_color='35de8c', fill_type='solid')

chart = ScatterChart()
chart.title = "Đồ thị F-a"
chart.style = 13
chart.x_axis.title = "Lực kéo F (N)"
chart.y_axis.title = "Gia tốc a (m/s)"
chart.legend = None

xvalues = Reference(ws, min_col=6, min_row=2, max_row=len(rows)-2)
yvalues = Reference(ws, min_col=7, min_row=2, max_row=len(rows)-2)
series = Series(xvalues, yvalues, title="")
chart.series.append(series)

ws.add_chart(chart, "I1")

wb.save("test//dothi.xlsx")