# thu vien gui
import PySimpleGUI as sg
# thu vien lay du lieu serial
import time
import serial
import serial.tools.list_ports
# thu vien chinh sua file excel
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import Workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
# thu vien he thong 
import os
import datetime

def xuatfiledothi(data, dir):
    wb = Workbook()
    ws = wb.active

    rows = [
        ["Lần thử", "Lực kéo (lt)", "Khối lượng", "Thời gian", "Quãng đường", "Gia tốc", "Lực kéo (tt)"],
    ]

    for d in data: 
        rows.append(d)

    for row in rows:
        ws.append(row)

    rowcount = len(rows)

    for cell in ws["1:1"]:
        cell.fill = PatternFill(start_color='35b1de', end_color='35b1de', fill_type='solid')
        cell.font = Font(bold=True)

    for rows in ws.iter_rows(min_row=2, min_col=1, max_col=7, max_row=len(rows)):
        for cell in rows:
            cell.fill = PatternFill(start_color='35de8c', end_color='35de8c', fill_type='solid')

    chart = ScatterChart()
    chart.title = "Đồ thị F-a"
    chart.style = 13
    chart.x_axis.title = "Lực kéo F (N)"
    chart.y_axis.title = "Gia tốc a (m/s)"
    chart.legend = None

    xvalues = Reference(ws, min_col=6, min_row=2, max_row=rowcount)
    yvalues = Reference(ws, min_col=7, min_row=2, max_row=rowcount)
    series = Series(xvalues, yvalues, title="")
    chart.series.append(series)

    ws.add_chart(chart, "I1")

    wb.save(dir)

def portselector():
    # hien thi tat ca cong serial dang mo tren may tinh
    ports = serial.tools.list_ports.comports()
    ports = sorted(ports)

    portlist = ""
    for i in range(len(ports)):
        port = ports[i].name
        desc = ports[i].description
        hwid = ports[i].hwid
        portlist += "{}. {}: {} [{}] \n".format(i+1, port, desc, hwid)
        # print("{}: {} [{}]".format(port, desc, hwid))

    # print(portlist)

    layout = [
        [sg.Text("Chọn cổng COM đến ESP32: ", background_color='#242526', text_color='#e7e9ed')],
        [sg.Text(portlist.strip(), background_color='#242526', text_color='#e7e9ed')], 
        [sg.InputText(background_color='#3a3b3c', text_color='#e7e9ed', border_width=0)], 
        [sg.Submit(button_text="Kết nối", button_color=('#3a3b3c', '#e7e9ed'))]
    ]
    win = sg.Window("Chọn cổng COM", layout, finalize=True, background_color='#242526', font=("Arial", 10))
    e, v = win.read()
    win.close()

    # chon cong com den esp32
    # i = int(input("Chọn cổng COM để kết nối đến ESP32: "))
    ser = serial.Serial(str(ports[int(v[0])-1].name), 9600, timeout=0.050)
    return ser, str(ports[int(v[0])-1].description)

def datain(ser, testcount, data):
    sout = ser.readline()
    sout_decoded = str(sout).split(";")
    print(sout_decoded)
    try:
        print(testcount, ". ", sout_decoded[2], "(ms)")
        testcount += 1
        while True:
            layout = [
                [sg.Text(f"Nhập dữ liệu còn thiếu của lần đo thứ {testcount}:",  background_color='#242526', text_color='#e7e9ed')],
                [sg.Text(f"Dữ liệu thời gian từ bộ đo: {sout_decoded[2]}(ms)",  background_color='#242526', text_color='#e7e9ed')],
                [sg.Text("Lực kéo (Lý thuyết): ",  background_color='#242526', text_color='#e7e9ed'), sg.InputText( background_color='#3a3b3c', text_color='#e7e9ed', border_width=0)],
                [sg.Text("Khối lượng:             ",  background_color='#242526', text_color='#e7e9ed'), sg.InputText( background_color='#3a3b3c', text_color='#e7e9ed', border_width=0)],
                [sg.Text("Quãng đường:         ",  background_color='#242526', text_color='#e7e9ed'), sg.InputText( background_color='#3a3b3c', text_color='#e7e9ed', border_width=0)],
                [sg.Submit(button_text="Hoàn tất",  button_color=('#3a3b3c', '#e7e9ed'))]
            ]
            win = sg.Window("Nhập dữ liệu đo", layout, finalize=True, background_color='#242526', font=('Arial', 15))
            e, v = win.read()
            win.close()
            if (v[0] != "" and v[1] != "" and v[2] != ""): break
            else: sg.Popup("Các ô dữ liệu không được để trống!", title="Chú ý", background_color='#242526', text_color='#e7e9ed', button_color=('#3a3b3c', '#e7e9ed'))
        time = float(sout_decoded[2]) / 1000 #ms to s
        acceleration = round(((float(v[2])*2)/(time*time)), 2)
        data.append([testcount, float(v[0]), float(v[1]), time, float(v[2]), acceleration, round((acceleration*float(v[1])),2)])
        print(data)
        return data, testcount
    except Exception as e:
        print("oops ", e)
        sg.Popup(e, title="Lỗi", background_color='#242526', text_color='#e7e9ed', button_color=('#3a3b3c', '#e7e9ed'))
        testcount -= 1
        return data, testcount

if __name__ == "__main__":
    # init serial port
    ser, ser_desc = portselector()
    print(ser.name, ser_desc)
    # bien dem so lan thu
    testcount = 0

    # du lieu cua bang trong gui
    data = []
    headings = ["Lần thử", "Lực kéo (lt)", "Khối lượng", "△t", "Quãng đường", "Gia tốc", "Lực kéo (tt)"]

    # tao cua so chuong trinh
    menu = [
        ['&Tệp', ['&Xuất đồ thị...', '&Thoát']],
        ['&Trợ giúp', ['&Thông tin']]
    ]
    layout = [
        [sg.Menu(menu, tearoff=False)],
        [sg.Table(
            values=data, 
            headings=headings, 
            key="-t-", 
            auto_size_columns=False,  
            def_col_width=10, 
            num_rows=10, 
            justification="center", 
            expand_x=True, 
            expand_y=True, 
            font=("Arial", 15, "bold"), 
            # header_background_color=(), header_text_color=(),
            header_relief=sg.RELIEF_SOLID,
            background_color='#242526', 
            text_color='#e7e9ed',
            sbar_trough_color='#3a3b3c', 
            sbar_background_color='#242526', 
            sbar_arrow_color='#3a3b3c', 
            sbar_frame_color='#242526', 
            sbar_relief=sg.RELIEF_FLAT
        )], 
        # [sg.Button("Xuất đồ thị .xlsx", key="-dt-",  button_color=('#3a3b3c', '#e7e9ed'))],
        [sg.StatusBar(f"Đã kết nối tới cổng {ser.name} ({ser_desc})", background_color='#242526', text_color='#e7e9ed', size=(100,1), pad=(0,0), relief=sg.RELIEF_FLAT, justification='right', visible=True,)]
    ]
    win = sg.Window(f"Kết quả đo", layout, resizable=True, finalize=True, background_color='#242526', size=(1200, 400))

    # event loop
    while True:
        if (ser.in_waiting != 0):
            data, testcount = datain(ser, testcount, data)
            print(data)
            win["-t-"].update(values=data)
        e, v = win.read(timeout=500)
        if e == sg.WIN_CLOSED or e == "Thoát":
            break
        if e == "Xuất đồ thị...":
            dir = str(os.getcwd())
            name = datetime.datetime.now()
            name = name.strftime("%d%m%y_%H%M%S") + ".xlsx"
            layout = [
                [
                    sg.Text("Chọn thư mục: ", background_color='#242526', text_color='#e7e9ed'), 
                    sg.Input(key="-IN2-" ,change_submits=True, default_text=dir, background_color='#3a3b3c', text_color='#e7e9ed', border_width=0), 
                    sg.FolderBrowse(key="-IN-", button_color=('#3a3b3c', '#e7e9ed'))
                ],
                [
                    sg.Button("Chọn", key="Submit", button_color=('#3a3b3c', '#e7e9ed'))
                ]
            ]
            exp_win = sg.Window("Xuất đồ thị", layout, finalize=True, background_color='#242526')
            while True:
                event, values = exp_win.read()
                if event == sg.WIN_CLOSED or event=="Exit":
                    break
                elif event == "Submit":
                    dir = values["-IN2-"]
                    dir = dir + f"/{name}"
                    xuatfiledothi(data, dir)
                    sg.Popup(f"Đã xuất {name} tại đường dẫn {dir}.", title="Hoàn tất", background_color='#242526', text_color='#e7e9ed', button_color=('#3a3b3c', '#e7e9ed'))
                    break
            exp_win.close()

    win.close()