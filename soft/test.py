# thu vien chinh sua file excel
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import Workbook
from openpyxl.chart import (
    legend,
    trendline,
    marker,
    ScatterChart,
    Reference,
    Series,
)

path = r"C:\Users\doquo\Documents\PlatformIO\Projects\newton2ndlaw-experiment\test\text.xlsx"

def export_to_excel(path, headers, data, fig1_title, fig1_headers, fig1_data, fig2_title, fig2_headers, fig2_data):

    # create a workbook
    wb = Workbook()
    ws = wb.active

    # main data
    rows = []
    rows.append(headers)
    for d in data: rows.append(d)

    # data for creating fig1
    r1 = []
    r1.append(fig1_headers)
    for d in fig1_data: r1.append(d)

    # data for creating fig2
    r2 = []
    r2.append(fig2_headers)
    for d in fig2_data: r2.append(d)

    
    for r in rows: ws.append(r)
    ws.append([])
    for r in r1: ws.append(r)
    ws.append([])
    for r in r2: ws.append(r)
    
    # create figure 1
    chart1 = ScatterChart()
    chart1.title = fig1_title
    chart1.style = 13
    chart1.x_axis.title = fig1_headers[0]
    chart1.y_axis.title = fig1_headers[1]
    chart1.legend = None

    x1value = Reference(ws, min_col=2, min_row=(len(rows) + 3), max_row=(len(rows) + 1)+len(r1))
    y1value = Reference(ws, min_col=1, min_row=(len(rows) + 3), max_row=(len(rows) + 1)+len(r1))
    series1 = Series(x1value, y1value)
    series1.marker= marker.Marker('circle')
    series1.graphicalProperties.line.noFill=True
    series1.trendline = trendline.Trendline(trendlineType='linear', dispEq=True)
    chart1.series.append(series1)

    # create figure 2
    chart2 = ScatterChart()
    chart2.title = fig2_title
    chart2.style = 13
    chart2.x_axis.title = fig2_headers[0]
    chart2.y_axis.title = fig2_headers[1]
    chart2.legend = None

    x2value = Reference(ws, min_col=2, min_row=(len(rows) + 2)+len(r1)+2, max_row=(len(rows) + 2)+len(r1)+len(r2))
    y2value = Reference(ws, min_col=1, min_row=(len(rows) + 2)+len(r1)+2, max_row=(len(rows) + 2)+len(r1)+len(r2))
    series2 = Series(x2value, y2value)
    series2.marker= marker.Marker('circle')
    series2.graphicalProperties.line.noFill=True
    series2.trendline = trendline.Trendline(trendlineType='linear', dispEq=True)
    chart2.series.append(series2)

    ws.add_chart(chart1, "I1")
    ws.add_chart(chart2, "I18")

    # tô màu
    for cell in ws["1:1"]:
        cell.fill = PatternFill(start_color='35b1de', end_color='35b1de', fill_type='solid')
        cell.font = Font(bold=True)
    
    for i in range(1, len(rows)):
        for cell in ws[f"{i+1}:{i+1}"]:
            cell.fill = PatternFill(start_color='c6efce', end_color='c6efce', fill_type='solid')
            cell.font = Font(bold=False)

    wb.save(path)


export_to_excel(path, ["Lần thử", "Lực kéo (lt)", "Khối lượng", "Thời gian", "Quãng đường", "Gia tốc"], [[1, 1.0, 0.3, 0.55, 0.5, 3.31], [2, 1.0, 0.4, 0.64, 0.5, 2.44], [3, 1.0, 0.5, 0.71, 0.5, 1.98], [4, 2.0, 0.5, 0.5, 0.5, 4.0], [5, 3.0, 0.5, 0.42, 0.5, 5.67]], "Đồ thị m+M (const)", ["Lực kéo", "Gia tốc"], [[1.0, 1.98], [2.0, 4.0], [3.0, 5.67]], "Đồ thị F (const)", ["1/m+M", "Gia tốc"], [[3.33, 3.31], [2.5, 2.44], [2.0, 1.98]])


exp_temp = {
        "name": "Định luật 2 newton",
        "desc": '''Tiến hành:
Bước 1: Lực kéo F có độ lớn tăng dần 1 N, 2 N và 3 N (bằng cách móc thêm các quả nặng vào đầu dây vắt qua ròng rọc). 
Bước 2: Ghi vào Bảng 15.1 độ lớn lực kéo F và tổng khối lượng của hệ (gồm xe trượt và các quả nặng đặt vào xe), ứng với mỗi lần thí nghiệm.
Bước 3: Do thời gian chuyển động của xe; đồng hồ bắt đầu đếm từ lúc tám chân sáng đi qua cổng quang điện 1 và kết thúc đêm khi tấm chắn vượt qua cổng quang điện 2.
Bước 4: Gia tốc a được tính từ công thức: a = v0*t + 1/2*a*t^2 (đặt xe trượt có gắn tấm chấn sáng sao cho tấm chắn này sát với cổng quang điện 1 để v = 0; s= 0,5 m là khoảng cách giữa hai cổng quang điện trong thí nghiệm). Đo thời gian túng với mỗi làn thí nghiệm.

Thảo luận:
a) Dựa vào số liệu trong Bảng 15.1, hãy vẽ đồ thị chỉ sự phụ thuộc của gia tốc a:
- Vào F (ứng với m + M = 0,5 kg), (Hình 15.3a). Đô thị có phải là đường tháng không? Tại sao?
- Vào 1 m+M (ứng với F - 1 N), đồ thị có phải là đường thẳng không? Tại sao?
b) Nếu kết luận về sự phụ thuộc của gia tốc vào độ lớn của lực tác dụng và khối lượng
của vật.''',
        "img": "assets//img1.png",
        "table_data": {
            "headers": ["Lần thử", "Lực kéo (lt)", "Khối lượng", "Thời gian", "Quãng đường", "Gia tốc"],
        },
        "plot": {
            "name": "Đồ thị m+M (const)",
            "x_name": "F (N)",
            "y_name": "a (m/s)",
            "headers": ["Lực kéo (F)", "Gia tốc (m/s)"]
        },
        "plot2": {
            "name": "Đồ thị F (const)",
            "x_name": "1/(m+M) (kg)",
            "y_name": "a (m/s)",
            "headers": ["1/m+M (kg)", "Gia tốc (m/s)"],
        },
    }